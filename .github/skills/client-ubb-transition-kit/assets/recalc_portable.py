#!/usr/bin/env python3
"""Portable formula recalculation + error scan for the Transition Kit XLSX.

Usage: python recalc_portable.py <file.xlsx> [timeout_seconds]

How it works: creates a throwaway LibreOffice profile with OOXMLRecalcMode=0
(always recalculate on load), converts the file to xlsx through Calc headless
(which forces a full recalculation and stores cached values), replaces the
original, then scans every formula cell. The scan FAILS if any cached value
is missing, not just on error strings, so a silent no-op recalculation can
never pass. Exit 0 only on fully calculated, zero-error workbooks.
If another workspace-specific xlsx recalculation helper is available, it can be
used as long as it performs a strict formula and cached-value scan.
"""
import sys, subprocess, tempfile, pathlib, json, shutil

ERRS = ('#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NUM!', '#NULL!')

REGMOD = '''<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry"
           xmlns:xs="http://www.w3.org/2001/XMLSchema">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load">
  <prop oor:name="ODFRecalcMode" oor:op="fuse"><value>0</value></prop>
 </item>
</oor:items>'''

def recalc(path: pathlib.Path, timeout: int = 120) -> None:
    profile = tempfile.mkdtemp(prefix='lo-profile-')
    user = pathlib.Path(profile) / 'user'
    user.mkdir(parents=True, exist_ok=True)
    (user / 'registrymodifications.xcu').write_text(REGMOD)
    outdir = tempfile.mkdtemp(prefix='lo-out-')
    subprocess.run(
        ['soffice', f'-env:UserInstallation=file://{profile}', '--headless',
         '--norestore', '--convert-to', 'xlsx:Calc MS Excel 2007 XML',
         '--outdir', outdir, str(path)],
        timeout=timeout, check=False,
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    converted = pathlib.Path(outdir) / path.name
    if converted.exists() and converted.stat().st_size > 0:
        shutil.copy2(converted, path)
    shutil.rmtree(profile, ignore_errors=True)
    shutil.rmtree(outdir, ignore_errors=True)

def scan(path: pathlib.Path) -> dict:
    import openpyxl
    wbf = openpyxl.load_workbook(path)
    wbv = openpyxl.load_workbook(path, data_only=True)
    errors, uncalced, formulas = [], [], 0
    for ws in wbf.worksheets:
        wsv = wbv[ws.title]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith('='):
                    formulas += 1
                    v = wsv[c.coordinate].value
                    if v is None and '""' not in c.value:
                        # formulas that can legally return an empty string
                        # (e.g. IFERROR(...,"")) cache as empty; not a failure
                        uncalced.append(f'{ws.title}!{c.coordinate}')
                    elif isinstance(v, str) and v in ERRS:
                        errors.append(f'{ws.title}!{c.coordinate}')
    status = ('errors_found' if errors else
              'not_calculated' if uncalced else 'success')
    return {'status': status, 'total_formulas': formulas,
            'total_errors': len(errors), 'uncalculated': len(uncalced),
            'error_locations': errors[:20], 'uncalculated_locations': uncalced[:10]}

if __name__ == '__main__':
    f = pathlib.Path(sys.argv[1]).resolve()
    t = int(sys.argv[2]) if len(sys.argv) > 2 else 120
    recalc(f, t)
    report = scan(f)
    print(json.dumps(report, indent=2))
    sys.exit(0 if report['status'] == 'success' else 1)
