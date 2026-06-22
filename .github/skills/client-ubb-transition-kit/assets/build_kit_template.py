"""
Client UBB Transition Kit, XLSX generator (BTG v11 master structure).
Fill CLIENT below and place client_monthly.csv and client_models.csv next to
this script (copy the BTG masters from examples/ and replace the data), then run.
ALWAYS recalc with scripts/recalc.py afterwards until zero errors.
"""
CLIENT = {
    "name": "Banco BTG Pactual S.A.",
    "tpid": 5336866,
    "plan": "Enterprise",
    "cb_seats": 1693,
    "ce_seats": 2438,
    "price_per_credit": 0.01,
    "monthly_consumption_usd": 231704,   # official Token Based Billing report
    "current_pru_overage_usd": 3699,
    "promo_months_covered": 2,           # months of the forecast FY inside the promo window
    "acd_band": 0.15,
    "azure_sub": "Yes",
    "growth_10m": 2.68,
    "cost_centers": 10,
    "report_label": "April 2026",
    "version": "v11_0_0",
    "date": "2026-06-12",
}

import pandas as pd, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

RED='F25022'; GREEN='7FBA00'; BLUE='00A4EF'; YELLOW='FFB900'
DARK='1A1A1A'; GRAY='F2F2F2'; LGRAY='FAFAFA'; WHITE='FFFFFF'
INPUT_FILL=PatternFill('solid', start_color='FFF2CC')   # yellow editable
HDR=lambda c: PatternFill('solid', start_color=c)
F=lambda **k: Font(name='Segoe UI', **k)
FIN_BLUE='0000FF'  # hardcoded input font
CUR='$#,##0;($#,##0);-'; CUR2='$#,##0.00;($#,##0.00);-'; PCT='0.0%'; NUM='#,##0;(#,##0);-'
thin=Border(bottom=Side(style='thin', color='DDDDDD'))
box=Border(*[Side(style='thin', color='BBBBBB')]*4)

# Input contract: two CSVs next to this script (see examples/ for the BTG masters)
#   client_monthly.csv : mes,ghcp_seats,ghcp_acr_usd,pru_acr_usd,pru_units,ghe_total_seats
#   client_models.csv  : modelo,custo_tokens_usd,usuarios_unicos,tokens_input,tokens_output,tokens_cache_criacao,tokens_cache_leitura
mensal = pd.read_csv('client_monthly.csv').reset_index(drop=True)
models = pd.read_csv('client_models.csv').sort_values('custo_tokens_usd', ascending=False).reset_index(drop=True)

wb = Workbook()

def tag(ws, label, color, title, sub):
    ws.sheet_view.showGridLines=False
    c=ws['B2']; c.value=label; c.font=F(bold=True,size=8,color=WHITE); c.fill=HDR(color)
    c.alignment=Alignment(horizontal='center', vertical='center')
    ws['B4']=title; ws['B4'].font=F(bold=True,size=18,color=DARK)
    ws['B5']=sub; ws['B5'].font=F(size=10,color='666666')

def footer(ws, row):
    ws.cell(row=row,column=2,value='Frontier Cockpit Team, Software Global Black Belt | frontier-cockpit@example.com | Microsoft').font=F(size=8,color='999999')

def th(ws,row,col,vals,color=DARK):
    for i,v in enumerate(vals):
        c=ws.cell(row=row,column=col+i,value=v)
        c.font=F(bold=True,size=9,color=WHITE); c.fill=HDR(color)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

def cellv(ws,r,c,v,nf=None,bold=False,fill=None,fontcolor=DARK,size=10):
    x=ws.cell(row=r,column=c,value=v)
    x.font=F(size=size,bold=bold,color=fontcolor)
    if nf: x.number_format=nf
    if fill: x.fill=fill
    return x

def colw(ws,widths,start=1):
    for i,w in enumerate(widths,start=start):
        ws.column_dimensions[get_column_letter(i)].width=w

status_dv = DataValidation(type='list', formula1='"Not started,In progress,Done,Blocked"', allow_blank=True)
plan_dv = DataValidation(type='list', formula1='"Business,Enterprise"', allow_blank=False)

# ============================================================= COVER
ws=wb.active; ws.title='Cover'
ws.sheet_view.showGridLines=False
for i,c in enumerate([RED,GREEN,BLUE,YELLOW]):
    ws.cell(row=2,column=2+i).fill=HDR(c)
    ws.row_dimensions[2].height=6
ws['B5']='GitHub Copilot UBB Transition Kit'; ws['B5'].font=F(bold=True,size=26,color=DARK)
ws['B6']=f"{CLIENT['name']} | TPID {CLIENT['tpid']} | {CLIENT['plan']}"; ws['B6'].font=F(size=13,color='444444')
ws['B8']='Dynamic budget engine, scenario forecasting, optimization program (R1-R7) and execution tracking for the Usage-Based Billing transition.'
ws['B8'].font=F(size=10,color='666666'); ws['B8'].alignment=Alignment(wrap_text=True)
ws['B10']='Frontier Cockpit Team, Software Global Black Belt'; ws['B10'].font=F(bold=True,size=11)
ws['B11']='frontier-cockpit@example.com'; ws['B11'].font=F(size=10,color='0563C1')
ws['B12']='Version 11.0.0 | 2026-06-12 | Confidential, internal account team use'; ws['B12'].font=F(size=9,color='999999')
ws['B14']='Companion documents (single source of truth set)'; ws['B14'].font=F(bold=True,size=11)
docs=[('Internal analysis HTML','BTG_Pactual_Analise_Budget_v8 (PT-BR, EN, ES)'),
      ('Client-facing HTML','BTG_Pactual_Analise_Budget_CLIENTE_v8 (PT-BR, EN, ES)'),
      ('Data backbone','db.json v2.0.0, Brazil UBB collection (tpid 5336866)'),
      ('This workbook','Engine, forecast, simulator, action plan and KPI tracking in English')]
for i,(a,b) in enumerate(docs):
    cellv(ws,15+i,2,a,bold=True,size=10); cellv(ws,15+i,4,b,size=10,fontcolor='555555')
ws['B21']='Workbook map'; ws['B21'].font=F(bold=True,size=11)
mapa=[('INPUT','Client Intake','everything yellow is editable; the entire kit recalculates'),
 ('ENGINE','Engine | Consumption FY26 | Model Mix','plan math, audited actuals from db.json, premium concentration'),
 ('BUSINESS CASE','Dashboard | Forecast FY27 | Simulator | ROI Scenarios','KPIs, September curve, what-if levers, A to D scenario ladder'),
 ('EXECUTION','Action Plan R1-R7 | 90-Day Plan | Monthly Checkpoints | KPI Tracker','actions with owner and status, weekly cadence, traffic lights'),
 ('GOVERNANCE','Governance | Maturity | Glossary | Sources','budgets and guardrails, maturity score, definitions, audit trail')]
th(ws,22,2,['Block','Sheets','What it does'])
for i,row in enumerate(mapa):
    for j,v in enumerate(row):
        cellv(ws,23+i,2+j,v,size=9,bold=(j==0))
        if i%2==0: ws.cell(row=23+i,column=2+j).fill=HDR(GRAY)
colw(ws,[2,18,46,70])
footer(ws,30)

# ============================================================= HOW TO USE
ws=wb.create_sheet('How to Use')
tag(ws,'GUIDE',BLUE,'How to Use This Kit','Five minutes to understand the flow. Conventions follow finance modeling standards.')
items=[
 ('1','Fill the Client Intake','Yellow cells only. Client, plan, seats, report consumption, ACD band. Everything downstream recalculates automatically.'),
 ('2','Read the Dashboard','KPIs, the three scenarios, the September curve and the model mix. This is the page you present.'),
 ('3','Pressure-test with the Simulator','Move the R1-R7 reduction levers. Combined effect is multiplicative, never additive, so it cannot exceed 100%.'),
 ('4','Commit to the plan','Action Plan R1-R7, 90-Day Plan and Monthly Checkpoints carry owner, due date and status dropdowns.'),
 ('5','Track weekly','KPI Tracker: enter actuals, variance and traffic lights are automatic. Governance holds the budget guardrails.'),
]
th(ws,7,2,['Step','What','How'])
for i,row in enumerate(items):
    for j,v in enumerate(row): cellv(ws,8+i,2+j,v,size=10,bold=(j==1))
    ws.cell(row=8+i,column=4).alignment=Alignment(wrap_text=True)
    ws.row_dimensions[8+i].height=28
ws['B15']='Color conventions'; ws['B15'].font=F(bold=True,size=11)
conv=[('Yellow fill','editable input, safe to change',INPUT_FILL,None),
      ('Blue font','hardcoded input value',None,FIN_BLUE),
      ('Black font','formula, do not overwrite',None,DARK),
      ('Green font','link from another sheet',None,'008000')]
for i,(a,b,fl,fc) in enumerate(conv):
    c=cellv(ws,16+i,2,a,size=10,bold=True,fontcolor=fc or DARK)
    if fl: c.fill=fl
    cellv(ws,16+i,3,b,size=10,fontcolor='555555')
colw(ws,[2,16,42,95])
footer(ws,22)

# ============================================================= CLIENT INTAKE
ws=wb.create_sheet('Client Intake')
tag(ws,'INPUT',YELLOW,'Client Intake','Yellow cells drive the entire kit. Audited values from the April 2026 GitHub billing report.')
fields=[
 ('Client name',CLIENT['name'],'legal entity', None,'s'),
 ('TPID',CLIENT['tpid'],'Microsoft Top Parent ID', NUM,'n'),
 ('Plan',CLIENT['plan'],'defines included credits and seat price', None,'s'),
 ('Business seats (CB)',CLIENT['cb_seats'],'from billing report Apr/2026', NUM,'n'),
 ('Enterprise seats (CE)',CLIENT['ce_seats'],'from billing report Apr/2026', NUM,'n'),
 ('Price per credit (USD)',CLIENT['price_per_credit'],'GitHub list price', CUR2,'n'),
 ('Monthly token consumption (USD)',CLIENT['monthly_consumption_usd'],'total token cost, Apr/2026 report; matches db.json model mix sum', CUR,'n'),
 ('Current monthly PRU overage (USD)',CLIENT['current_pru_overage_usd'],'old model: spend above included PRU, Apr/2026', CUR,'n'),
 ('Promo window (months covered)',CLIENT['promo_months_covered'],'promo window; months of FY27 covered (Jul, Aug)', NUM,'n'),
 ('Current ACD band (%)',CLIENT['acd_band'],'the Azure discount the new deal must beat', PCT,'n'),
 ('Azure subscription available',CLIENT['azure_sub'],'required for the above-table commercial path', None,'s'),
 ('Consumption growth, 10 months',CLIENT['growth_10m'],'computed from the series; see Consumption FY26 for the computed series', '0%','n'),
 ('Cost centers',CLIENT['cost_centers'],'used by Governance budget split', NUM,'n'),
]
th(ws,7,2,['Field','Value','Note'])
for i,(a,v,nt,nf,kind) in enumerate(fields):
    r=8+i
    cellv(ws,r,2,a,size=10,bold=True)
    c=cellv(ws,r,3,v,nf=nf,size=10,fontcolor=FIN_BLUE); c.fill=INPUT_FILL
    cellv(ws,r,4,nt,size=9,fontcolor='777777')
ws.add_data_validation(plan_dv); plan_dv.add('C10')
colw(ws,[2,34,22,80])
footer(ws,23)
IN={'client':'C8','tpid':'C9','plan':'C10','cb':'C11','ce':'C12','ppc':'C13','cons':'C14','pru_over':'C15','promo_m':'C16','acd':'C17','growth':'C19','cc':'C20'}

# ============================================================= ENGINE
ws=wb.create_sheet('Engine')
tag(ws,'ENGINE',RED,'Budget Engine','Plan table, pools, overage and the three scenarios. All formulas, zero hardcoded results.')
ws['B7']='Plan table (GitHub Copilot list)'; ws['B7'].font=F(bold=True,size=11)
th(ws,8,2,['Plan','Included std (credits/seat)','Included promo (credits/seat)','Price per seat (USD)'])
for i,(p,a,b,c) in enumerate([('Business',1900,3000,19),('Enterprise',3900,7000,39)]):
    r=9+i
    cellv(ws,r,2,p,size=10,bold=True)
    cellv(ws,r,3,a,NUM,fontcolor=FIN_BLUE); cellv(ws,r,4,b,NUM,fontcolor=FIN_BLUE); cellv(ws,r,5,c,CUR,fontcolor=FIN_BLUE)
ws['B12']='Derived (do not edit)'; ws['B12'].font=F(bold=True,size=11)
derived=[
 ('License cost (USD/month)', f"='Client Intake'!{IN['cb']}*E9+'Client Intake'!{IN['ce']}*E10", CUR),
 ('Included pool, standard (USD/month)', f"=('Client Intake'!{IN['cb']}*C9+'Client Intake'!{IN['ce']}*C10)*'Client Intake'!{IN['ppc']}", CUR),
 ('Included pool, promo (USD/month)', f"=('Client Intake'!{IN['cb']}*D9+'Client Intake'!{IN['ce']}*D10)*'Client Intake'!{IN['ppc']}", CUR),
 ('Overage, standard (USD/month)', f"=MAX(0,'Client Intake'!{IN['cons']}-C14)", CUR),
 ('Overage, promo (USD/month)', f"=MAX(0,'Client Intake'!{IN['cons']}-C15)", CUR),
 ('Consumption / included ratio', f"=IF(C14=0,0,'Client Intake'!{IN['cons']}/C14)", '0.00"x"'),
 ('Pool coverage of consumption', f"=IF('Client Intake'!{IN['cons']}=0,0,MIN(1,C14/'Client Intake'!{IN['cons']}))", PCT),
]
for i,(a,f,nf) in enumerate(derived):
    r=13+i
    cellv(ws,r,2,a,size=10); cellv(ws,r,3,f,nf,bold=True)
ws['B21']='The three scenarios'; ws['B21'].font=F(bold=True,size=11)
th(ws,22,2,['Scenario','USD/month','USD/year','Delta vs today (USD/mo)','Delta vs today (%)'])
scen=[
 ('Today (PRU model)', f"=C13+'Client Intake'!{IN['pru_over']}"),
 ('No action (Sep 2026 onward)', '=C13+C16'),
 ('With promo (Jun-Aug 2026)', '=C13+C17'),
]
for i,(a,f) in enumerate(scen):
    r=23+i
    cellv(ws,r,2,a,size=10,bold=True)
    cellv(ws,r,3,f,CUR,bold=True)
    cellv(ws,r,4,f'=C{r}*12',CUR)
    cellv(ws,r,5,f'=C{r}-$C$23',CUR)
    cellv(ws,r,6,f'=IF($C$23=0,0,(C{r}-$C$23)/$C$23)',PCT)
cellv(ws,26,2,'September jump (promo expiry)',size=10,bold=True)
cellv(ws,26,3,'=C24-C25',CUR,bold=True,fontcolor=RED)
cellv(ws,26,4,'=(C24-C25)*12',CUR)
cellv(ws,27,2,'Annual exposure above today, no action',size=10)
cellv(ws,27,3,'=(C24-C23)*12',CUR,bold=True)
colw(ws,[2,40,22,22,24,22])
footer(ws,29)

# ============================================================= CONSUMPTION FY26
ws=wb.create_sheet('Consumption FY26')
tag(ws,'DATA',GREEN,'Consumption FY26, Audited Series','Monthly actuals from db.json v2.0.0 (MSX FY26). Growth columns are formulas.')
th(ws,7,2,['Month','GHCP seats','GHCP ACR (USD)','PRU (USD)','PRU units','GHE seats','ACR MoM %','ACR index (Jul=100)'])
r0=8
for i,(_,r) in enumerate(mensal.iterrows()):
    ri=r0+i
    cellv(ws,ri,2,r['mes'],size=10)
    cellv(ws,ri,3,round(float(r['ghcp_seats']),1),NUM,fontcolor=FIN_BLUE)
    cellv(ws,ri,4,round(float(r['ghcp_acr_usd']),0),CUR,fontcolor=FIN_BLUE)
    cellv(ws,ri,5,round(float(r['pru_acr_usd']),0),CUR,fontcolor=FIN_BLUE)
    cellv(ws,ri,6,round(float(r['pru_units']),0),NUM,fontcolor=FIN_BLUE)
    cellv(ws,ri,7,round(float(r['ghe_total_seats']),0),NUM,fontcolor=FIN_BLUE)
    cellv(ws,ri,8,f'=IF(D{ri-1}=0,0,D{ri}/D{ri-1}-1)' if i>0 else None,PCT)
    cellv(ws,ri,9,f'=D{ri}/$D${r0}*100','#,##0')
n=len(mensal); last=r0+n-1
sumr=last+2
cellv(ws,sumr,2,'Computed signals',bold=True,size=11)
sig=[('Growth over the series (GHCP ACR)',f'=D{last}/D{r0}-1',PCT),
     ('Average MoM growth, last 4 months',f'=AVERAGE(H{last-3}:H{last})',PCT),
     ('Seats added in the series',f'=C{last}-C{r0}',NUM),
     ('PRU spend multiplied by',f'=IF(E{r0}=0,0,E{last}/E{r0})','0"x"')]
for i,(a,f,nf) in enumerate(sig):
    cellv(ws,sumr+1+i,2,a,size=10); cellv(ws,sumr+1+i,3,f,nf,bold=True)
colw(ws,[2,14,13,16,14,12,12,12,16])
footer(ws,sumr+7)

# ============================================================= MODEL MIX
ws=wb.create_sheet('Model Mix')
tag(ws,'DATA',YELLOW,'Model Mix, April 2026','Token cost by AI model from db.json (Model Level Details). Shares and concentration are formulas.')
th(ws,7,2,['Model','Token cost (USD)','Share of bill','Unique users','Input tokens (M)','Output tokens (M)','Cache read tokens (M)','Cache read share'])
r0=8; nm=len(models)
for i,(_,r) in enumerate(models.iterrows()):
    ri=r0+i
    cellv(ws,ri,2,r['modelo'],size=10,bold=(i<3))
    cellv(ws,ri,3,round(float(r['custo_tokens_usd']),0),CUR,fontcolor=FIN_BLUE)
    cellv(ws,ri,4,f'=C{ri}/SUM($C${r0}:$C${r0+nm-1})',PCT)
    cellv(ws,ri,5,int(r['usuarios_unicos']) if pd.notna(r['usuarios_unicos']) else None,NUM,fontcolor=FIN_BLUE)
    for j,k in enumerate(['tokens_input','tokens_output','tokens_cache_leitura']):
        cellv(ws,ri,6+j,round(float(r[k])/1e6,0) if pd.notna(r[k]) else 0,NUM,fontcolor=FIN_BLUE)
    cellv(ws,ri,9,f'=IF(F{ri}+G{ri}+H{ri}=0,0,H{ri}/(F{ri}+G{ri}+H{ri}))',PCT)
last=r0+nm-1; sr=last+2
cellv(ws,sr,2,'Total',bold=True); cellv(ws,sr,3,f'=SUM(C{r0}:C{last})',CUR,bold=True)
cellv(ws,sr,9,f'=SUMPRODUCT(H{r0}:H{last})/SUMPRODUCT(F{r0}:F{last}+G{r0}:G{last}+H{r0}:H{last})',PCT,bold=True)
cellv(ws,sr+1,2,'Claude family share',size=10)
cellv(ws,sr+1,3,f'=SUMIF(B{r0}:B{last},"*claude*",C{r0}:C{last})/C{sr}',PCT,bold=True)
cellv(ws,sr+2,2,'Top 2 models share',size=10)
cellv(ws,sr+2,3,f'=(LARGE(C{r0}:C{last},1)+LARGE(C{r0}:C{last},2))/C{sr}',PCT,bold=True)
cellv(ws,sr+3,2,'Why it matters',size=10,bold=True)
cellv(ws,sr+3,3,'Cache reads dominate token volume. Stable prefixes and context curation (R4, R5) attack exactly this line.',size=9,fontcolor='666666')
colw(ws,[2,24,16,12,12,14,14,18,14])
footer(ws,sr+5)
MM={'r0':r0,'last':last,'total_row':sr}

# ============================================================= SIMULATOR
ws=wb.create_sheet('Simulator')
tag(ws,'WHAT-IF',BLUE,'Budget Simulator, R1-R7 Levers','Editable reduction per lever. Combined effect is multiplicative: 1-(1-r1)x...x(1-r7). It can never exceed 100%.')
th(ws,7,2,['Lever','Reduction (editable)','Attacks','Official band','Note'])
levers=[
 ('R1 Output control',0.06,'output','40-70% of output','output costs ~4-5x input; instruction blocks cut verbosity'),
 ('R2 Model routing',0.10,'whole bill','40-70%','price per token spans two orders of magnitude; route by task'),
 ('R3 Local models (BYOK)',0.02,'routine tasks','single digit to ~15%','BYOK local models do not consume credits; not for completions'),
 ('R4 Context scoping',0.08,'input','40-80% of input','context is often 60%+ of input per turn'),
 ('R5 Cache and memory',0.06,'input','30-50% of input','stable prefixes are not reprocessed; cache reads are cheaper'),
 ('R6 Repository primitives',0.03,'input and bill','compound','versioned instructions, agents and prompts'),
 ('R7 Budgets and measurement',0.00,'variance','guardrail','does not cut the run rate; prevents runaway loops'),
]
r0=8
for i,(a,v,b,c,d) in enumerate(levers):
    ri=r0+i
    cellv(ws,ri,2,a,size=10,bold=True)
    x=cellv(ws,ri,3,v,PCT,fontcolor=FIN_BLUE); x.fill=INPUT_FILL
    cellv(ws,ri,4,b,size=9); cellv(ws,ri,5,c,size=9); cellv(ws,ri,6,d,size=9,fontcolor='777777')
last=r0+6
ws['B17']='Simulated result'; ws['B17'].font=F(bold=True,size=11)
res=[
 ('Combined consumption reduction', '=1-(1-C8)*(1-C9)*(1-C10)*(1-C11)*(1-C12)*(1-C13)*(1-C14)', PCT, True),
 ('Mature overage with program (USD/month)', f"=MAX(0,'Client Intake'!{IN['cons']}*(1-C18)-Engine!C14)", CUR, True),
 ('Mature monthly saving (USD)', '=Engine!C16-C19', CUR, False),
 ('Mature run-rate saving (USD/year)', '=C20*12', CUR, True),
 ('Overage drop at maturity', '=IF(Engine!C16=0,0,(Engine!C16-C19)/Engine!C16)', PCT, False),
 ("FY27 program savings (maturity curve)", "='Forecast FY27'!F22", CUR, True),
]
for i,(a,f,nf,b) in enumerate(res):
    r=18+i
    cellv(ws,r,2,a,size=10); cellv(ws,r,3,f,nf,bold=b)
cellv(ws,25,2,'Reference bands: a lean start delivers 20-30% reduction; a mature program 55-70%. Defaults above are lean. Sources sheet holds the references.',size=9,fontcolor='777777')
colw(ws,[2,30,16,16,18,62])
footer(ws,27)

# ============================================================= FORECAST FY27
ws=wb.create_sheet('Forecast FY27')
tag(ws,'BUSINESS CASE',RED,'Forecast FY27, The September Curve','Promo covers Jul-Aug. Ramp (yellow) is the program maturation; everything else is formula.')
th(ws,7,2,['Month','Ramp','No-action overage (USD)','With-program overage (USD)','Monthly saving (USD)','Cumulative saving (USD)','Scenario cost no action (USD)','Scenario cost with program (USD)'])
months=['Jul-26','Aug-26','Sep-26','Oct-26','Nov-26','Dec-26','Jan-27','Feb-27','Mar-27','Apr-27','May-27','Jun-27']
ramps=[0,0.05,0.15,0.30,0.45,0.60,0.75,0.85,0.92,0.96,0.98,1.00]
r0=10
ws.cell(row=9,column=2,value='FY27 (Microsoft fiscal year)').font=F(size=9,color='777777')
for i,(m,rp) in enumerate(zip(months,ramps)):
    ri=r0+i
    cellv(ws,ri,2,m,size=10)
    x=cellv(ws,ri,3,rp,PCT,fontcolor=FIN_BLUE); x.fill=INPUT_FILL
    promo_flag=f"ROW()-{r0}<'Client Intake'!{IN['promo_m']}+0"  # i < promo months
    cellv(ws,ri,4,f"=IF(ROW()-{r0}<'Client Intake'!{IN['promo_m']},Engine!$C$17,Engine!$C$16)",CUR)
    cellv(ws,ri,5,f"=MAX(0,'Client Intake'!$C$14*(1-C{ri}*Simulator!$C$18)-IF(ROW()-{r0}<'Client Intake'!$C$16,Engine!$C$15,Engine!$C$14))",CUR)
    cellv(ws,ri,6,f'=D{ri}-E{ri}',CUR)
    cellv(ws,ri,7,f'=SUM($F${r0}:F{ri})',CUR)
    cellv(ws,ri,8,f'=Engine!$C$13+D{ri}',CUR)
    cellv(ws,ri,9,f'=Engine!$C$13+E{ri}',CUR)
last=r0+11; tr=last+1
cellv(ws,tr,2,'Total FY27',bold=True)
for col,letter in [(4,'D'),(5,'E'),(6,'F'),(8,'H'),(9,'I')]:
    cellv(ws,tr,col,f'=SUM({letter}{r0}:{letter}{last})',CUR,bold=True)
cellv(ws,tr+2,2,'Reading',bold=True,size=10)
cellv(ws,tr+2,3,'No-action plateau lands in September. The with-program line bends as R1-R7 mature. Savings accumulate without costing Microsoft a dollar.',size=9,fontcolor='777777')
colw(ws,[2,10,9,20,22,18,20,24,26])
footer(ws,tr+4)
FC={'r0':r0,'last':last,'tr':tr}

# ============================================================= ROI SCENARIOS
ws=wb.create_sheet('ROI Scenarios')
tag(ws,'BUSINESS CASE',GREEN,'ROI, The Four FY27 Scenarios','The savings ladder. P3 percentages are editable; everything else recalculates.')
cellv(ws,7,2,'P3 discount inputs',bold=True,size=11)
p3=[('P3 published discount',0.15,'public table, does not beat the ACD'),
    ('P3 via Deal Desk',0.20,'subject to eligibility and approval'),]
for i,(a,v,nt) in enumerate(p3):
    r=8+i
    cellv(ws,r,2,a,size=10)
    x=cellv(ws,r,3,v,PCT,fontcolor=FIN_BLUE); x.fill=INPUT_FILL
    cellv(ws,r,4,nt,size=9,fontcolor='777777')
th(ws,11,2,['Scenario','FY27 total cost (USD)','Savings vs A (USD)','What it is'])
scen=[
 ('A No discount, no action', f"='Forecast FY27'!H22", '=0', 'baseline, full overage, no program'),
 ('B Promo + P3 15%', f"='Forecast FY27'!H22-'Forecast FY27'!D22*C8", '=C12-C13', 'published discount on token overage'),
 ('C Promo + P3 20% (Deal Desk)', f"='Forecast FY27'!H22-'Forecast FY27'!D22*C9", '=C12-C14', 'beats the ACD; needs approval'),
 ('D P3 20% + optimization program', f"='Forecast FY27'!I22-'Forecast FY27'!E22*C9", '=C12-C15', 'recommended path: discount on the reduced overage plus R1-R7 savings'),
]
for i,(a,f1,f2,nt) in enumerate(scen):
    r=12+i
    cellv(ws,r,2,a,size=10,bold=True)
    cellv(ws,r,3,f1,CUR,bold=True)
    cellv(ws,r,4,f2,CUR)
    cellv(ws,r,5,nt,size=9,fontcolor='777777')
cellv(ws,17,2,'Savings decomposition, scenario D',bold=True,size=11)
dec=[('Commercial discount component',"='Forecast FY27'!E22*C9",'the P3 cut on the optimized overage'),
     ('Efficiency redemption component',"='Forecast FY27'!F22",'R1-R7 savings; costs Microsoft nothing'),
     ('Total scenario D savings','=C18+C19','redemption is worth more than the discount')]
for i,(a,f,nt) in enumerate(dec):
    r=18+i
    cellv(ws,r,2,a,size=10); cellv(ws,r,3,f,CUR,bold=(i==2)); cellv(ws,r,4,nt,size=9,fontcolor='777777')
colw(ws,[2,38,22,20,66])
footer(ws,23)

# ============================================================= ACTION PLAN
ws=wb.create_sheet('Action Plan R1-R7')
tag(ws,'EXECUTION',RED,'Action Plan, R1-R7','Owner, due date and status per action. The Dashboard counts these statuses live.')
th(ws,7,2,['ID','Lever','Action','Owner','Due','Effort','Expected impact','Status','Notes'])
actions=[
 ('A01','R7 Budgets','Enable budgets and alerts per org and cost center (alert 80%, block 100%)','Client platform team','2026-07-15','Low','variance guardrail','In progress',''),
 ('A02','R2 Routing','Publish model allowlist and task-to-model policy (light-first)','GBB + client arch','2026-07-31','Medium','frontier share down','Not started',''),
 ('A03','R1 Output','Roll out instruction blocks limiting output verbosity in top repos','Client dev leads','2026-08-15','Low','output tokens down','Not started',''),
 ('A04','R4 Context','Scope context: #-mentions discipline, trim attachments, kill mega-threads','Client dev leads','2026-08-15','Medium','input tokens down','Not started',''),
 ('A05','R5 Cache','Stabilize system prompts and shared prefixes to raise cache hits','Client platform team','2026-08-31','Medium','cache read share up','Not started',''),
 ('A06','R6 Primitives','Version prompts, instructions and agent definitions in the repo','GBB + client arch','2026-09-15','Medium','compound savings','Not started',''),
 ('A07','R3 BYOK','Pilot local models for routine, non-completion tasks','Client arch','2026-09-30','High','routine offload','Not started',''),
 ('A08','Commercial','Size the P3 and open the Deal Desk case before promo expiry','MSFT account team','2026-08-15','Medium','beats the ACD','In progress',''),
 ('A09','Governance','Stand up the FinOps dashboard and weekly mix review','Client FinOps','2026-08-01','Medium','visibility','Not started',''),
 ('A10','Checkpoint','September review package: real overage vs forecast, milestones','GBB','2026-09-01','Low','decision gate','Not started',''),
]
r0=8
for i,row in enumerate(actions):
    for j,v in enumerate(row): cellv(ws,r0+i,2+j,v,size=9,bold=(j==0))
    ws.cell(row=r0+i,column=4).alignment=Alignment(wrap_text=True)
    ws.row_dimensions[r0+i].height=26
last=r0+len(actions)-1
ws.add_data_validation(status_dv); status_dv.add(f'I{r0}:I{last}')
for st,colr in [('Done','C6EFCE'),('In progress','FFF2CC'),('Blocked','F8CBAD'),('Not started','EFEFEF')]:
    ws.conditional_formatting.add(f'I{r0}:I{last}', CellIsRule(operator='equal', formula=[f'"{st}"'], fill=PatternFill('solid', start_color=colr)))
colw(ws,[2,7,14,58,22,12,10,20,14,24])
footer(ws,last+2)
AP={'r0':r0,'last':last}

# ============================================================= 90 DAY PLAN
ws=wb.create_sheet('90-Day Plan')
tag(ws,'EXECUTION',BLUE,'90-Day Plan, Weekly Cadence','Twelve weeks aligned to the transition. Status dropdowns feed the Dashboard.')
th(ws,7,2,['Week','Phase','KPI / focus','Main action','Target','Status'])
weeks=[
 ('W1','Baseline','Usage projection done','Ingest consumption, read the curve','projection complete'),
 ('W2','Baseline','Model allowlist defined','Publish task-to-model policy','allowlist live'),
 ('W3','Baseline','Team enabled on the change','Enablement session, pick pilot squad','pilot defined'),
 ('W4','Baseline','Per-user budget set','Configure user caps in the pilot','budget live in pilot'),
 ('W5','Instrument','Caching on long prompts','Enable caching, mark system prompts','caching in production'),
 ('W6','Instrument','FinOps dashboard live','Stand up billing overview and alerts','dashboard active'),
 ('W7','Instrument','Model mix under review','Weekly mix review; contain frontier','frontier share falling'),
 ('W8','Instrument','Budgets refined','Recalibrate with 30 days of real data','budgets calibrated'),
 ('W9','Optimize','Primitives published','Standardize prompts in the repo','primitives in use'),
 ('W10','Optimize','Agents scoped and capped','Configure scope, tools and limits','agents governed'),
 ('W11','Optimize','Redemption curve visible','Compare real vs forecast overage','redemption on track'),
 ('W12','Scale','September review ready','Consolidate milestones for checkpoint','review package ready'),
]
r0=8
for i,row in enumerate(weeks):
    for j,v in enumerate(row): cellv(ws,r0+i,2+j,v,size=9,bold=(j==0))
    cellv(ws,r0+i,7,'Not started',size=9)
last=r0+11
status_dv2 = DataValidation(type='list', formula1='"Not started,In progress,Done,Blocked"', allow_blank=True)
ws.add_data_validation(status_dv2); status_dv2.add(f'G{r0}:G{last}')
for st,colr in [('Done','C6EFCE'),('In progress','FFF2CC'),('Blocked','F8CBAD'),('Not started','EFEFEF')]:
    ws.conditional_formatting.add(f'G{r0}:G{last}', CellIsRule(operator='equal', formula=[f'"{st}"'], fill=PatternFill('solid', start_color=colr)))
colw(ws,[2,7,12,30,42,26,14])
footer(ws,last+2)
WP={'r0':r0,'last':last}

# ============================================================= MONTHLY CHECKPOINTS
ws=wb.create_sheet('Monthly Checkpoints')
tag(ws,'EXECUTION',GREEN,'Monthly Checkpoints','Six months. September is the lock: promo expires, real overage shows up.')
th(ws,7,2,['Month','Checkpoint','Primary indicator','Target','Status'])
cps=[
 ('Jul-26','Promo active, baseline measured','overage covered by promo','clear baseline'),
 ('Aug-26','Instrumentation live','caching and dashboard active','foundation ready'),
 ('Sep-26','REVIEW, promo expires','real overage appears; milestones met?','extension decided'),
 ('Oct-26','Redemption maturing','overage falling vs plateau','descending trajectory'),
 ('Nov-26','Continuous optimization','model mix under control','frontier share lower'),
 ('Dec-26','Halfway point','redemption ~60% mature','savings accumulating'),
]
r0=8
for i,row in enumerate(cps):
    for j,v in enumerate(row): cellv(ws,r0+i,2+j,v,size=9,bold=(j==0 or 'REVIEW' in str(row[1]) and j==1))
    cellv(ws,r0+i,6,'Not started',size=9)
last=r0+5
status_dv3 = DataValidation(type='list', formula1='"Not started,In progress,Done,Blocked"', allow_blank=True)
ws.add_data_validation(status_dv3); status_dv3.add(f'F{r0}:F{last}')
for st,colr in [('Done','C6EFCE'),('In progress','FFF2CC'),('Blocked','F8CBAD'),('Not started','EFEFEF')]:
    ws.conditional_formatting.add(f'F{r0}:F{last}', CellIsRule(operator='equal', formula=[f'"{st}"'], fill=PatternFill('solid', start_color=colr)))
ws.conditional_formatting.add(f'B{r0+2}:F{r0+2}', FormulaRule(formula=['TRUE'], fill=PatternFill('solid', start_color='FDEFEA')))
colw(ws,[2,9,34,38,26,14])
footer(ws,last+2)
CP={'r0':r0,'last':last}

# ============================================================= KPI TRACKER
ws=wb.create_sheet('KPI Tracker')
tag(ws,'EXECUTION',YELLOW,'KPI Tracker, Enter Actuals Weekly','Yellow cells take real data. Variance and traffic lights are automatic.')
th(ws,7,2,['KPI','Unit','Target','W1','W2','W3','W4','W5','W6','W7','W8','Latest','Variance vs target'])
kpis=[
 ('Monthly overage run-rate','USD/month','=Simulator!C19',CUR),
 ('Frontier model share of bill','%',0.10,PCT),
 ('Cache read share of tokens','%',0.85,PCT),
 ('Seats over per-user budget','count',0,NUM),
 ('Weekly active users (WAU)','%',0.65,PCT),
 ('Actions done (Action Plan)','count',f"=COUNTIF('Action Plan R1-R7'!I{AP['r0']}:I{AP['last']},\"Done\")",NUM),
]
r0=8
for i,(a,u,t,nf) in enumerate(kpis):
    ri=r0+i
    cellv(ws,ri,2,a,size=10,bold=True); cellv(ws,ri,3,u,size=9,fontcolor='777777')
    cellv(ws,ri,4,t,nf)
    for w in range(8):
        x=cellv(ws,ri,5+w,None,nf); x.fill=INPUT_FILL
    cellv(ws,ri,13,f'=IFERROR(LOOKUP(2,1/(E{ri}:L{ri}<>""),E{ri}:L{ri}),"")',nf,bold=True)
    cellv(ws,ri,14,f'=IF(M{ri}="","",M{ri}-D{ri})',nf)
last=r0+len(kpis)-1
ws.conditional_formatting.add(f'N{r0}:N{last}', ColorScaleRule(start_type='min', start_color='7FBA00', mid_type='num', mid_value=0, mid_color='FFFFFF', end_type='max', end_color='F25022'))
cellv(ws,last+2,2,'Note: for the overage and frontier KPIs, below target is good. Read the variance sign accordingly.',size=9,fontcolor='777777')
colw(ws,[2,30,10,12]+[9]*8+[12,16])
footer(ws,last+4)
KT={'r0':r0,'last':last}

# ============================================================= GOVERNANCE
ws=wb.create_sheet('Governance')
tag(ws,'GOVERNANCE',BLUE,'Budgets and Guardrails','Caps computed from the engine. Rule of thumb: alert at 80%, block at 100%.')
gov=[
 ('Mature target overage (USD/month)','=Simulator!C19',CUR),
 ('Licensed seats',f"='Client Intake'!{IN['cb']}+'Client Intake'!{IN['ce']}",NUM),
 ('Cap per user (USD/month)','=IF(C9=0,0,C8/C9)',CUR2),
 ('Cost centers',f"='Client Intake'!{IN['cc']}",NUM),
 ('Cap per cost center (USD/month)','=IF(C11=0,0,C8/C11)',CUR),
 ('Alert margin',0.8,PCT),
 ('Per-user alert trigger (USD)','=C10*C13',CUR2),
 ('Per-cost-center alert trigger (USD)','=C12*C13',CUR),
]
r0=8
for i,(a,f,nf) in enumerate(gov):
    r=r0+i
    cellv(ws,r,2,a,size=10)
    if a=='Alert margin':
        x=cellv(ws,r,3,f,nf,fontcolor=FIN_BLUE); x.fill=INPUT_FILL
    else:
        cellv(ws,r,3,f,nf,bold=True)
cellv(ws,17,2,'Three-layer metrics model',bold=True,size=11)
th(ws,18,2,['Layer','The question it answers','Primary owner'])
layers=[('1 Cost and governance','Is the financial impact under control? Overage trajectory, mix, budgets','GBB + client data'),
 ('2 Platform and architecture','Is the foundation being built? Primitives, caching, agents governed','Partners + GBB'),
 ('3 Organizational outcome','Did real value show up? Faster, safer delivery at the org level','All three sides')]
for i,row in enumerate(layers):
    for j,v in enumerate(row): cellv(ws,19+i,2+j,v,size=9,bold=(j==0))
cellv(ws,23,2,'Trap to avoid: the AI productivity paradox. Individual output rises while org throughput stalls. Measure layer 3, not anecdotes.',size=9,fontcolor='777777')
colw(ws,[2,38,18,60])
footer(ws,25)

# ============================================================= MATURITY
ws=wb.create_sheet('Maturity')
tag(ws,'DIAGNOSTIC',GREEN,'Maturity Assessment','Score each dimension 1-5. Stage and color are automatic.')
th(ws,7,2,['Dimension','Score (1-5)','What to assess'])
dims=[('Model policy and routing',3,'governance over which model per task'),
 ('Context curation and cache',3,'lean context discipline and caching'),
 ('Repository primitives',3,'standardized, versioned prompts and instructions'),
 ('Agent architecture',2,'agents with scope, tools and caps'),
 ('FinOps and observability',2,'cost dashboards, alerts and budgets live')]
r0=8
for i,(a,v,nt) in enumerate(dims):
    ri=r0+i
    cellv(ws,ri,2,a,size=10,bold=True)
    x=cellv(ws,ri,3,v,NUM,fontcolor=FIN_BLUE); x.fill=INPUT_FILL
    cellv(ws,ri,4,nt,size=9,fontcolor='777777')
last=r0+4
cellv(ws,last+2,2,'Total score',bold=True); cellv(ws,last+2,3,f'=SUM(C{r0}:C{last})',NUM,bold=True)
cellv(ws,last+3,2,'Stage',bold=True)
cellv(ws,last+3,3,f'=IF(C{last+2}<=10,"Initial",IF(C{last+2}<=18,"Evolving","Mature"))',bold=True)
ws.conditional_formatting.add(f'C{last+3}', CellIsRule(operator='equal', formula=['"Initial"'], fill=PatternFill('solid', start_color='F8CBAD')))
ws.conditional_formatting.add(f'C{last+3}', CellIsRule(operator='equal', formula=['"Evolving"'], fill=PatternFill('solid', start_color='FFF2CC')))
ws.conditional_formatting.add(f'C{last+3}', CellIsRule(operator='equal', formula=['"Mature"'], fill=PatternFill('solid', start_color='C6EFCE')))
colw(ws,[2,32,12,64])
footer(ws,last+6)

# ============================================================= GLOSSARY
ws=wb.create_sheet('Glossary')
tag(ws,'REFERENCE',YELLOW,'Glossary','The terms that show up in the conversation, in plain language.')
gl=[
 ('UBB','Usage-Based Billing: AI consumption above the included pool is metered and billed'),
 ('Credit','the billing unit; 1 credit = USD 0.01 at list'),
 ('Included pool','credits bundled per seat, pooled at the org level (Business 1,900; Enterprise 3,900 per seat per month)'),
 ('Promo pool','the 90-day promotion: Business 3,000 and Enterprise 7,000 credits per seat, Jun-Aug 2026'),
 ('Overage','consumption above the pool, billed at price per credit'),
 ('Pooling','heavy and light users share the same org pool; no credits expire individually'),
 ('PRU','Premium Request Units, the old metering model with an opaque cap'),
 ('P3','Pre-Purchase Plan: committed credit purchase with discount band'),
 ('ACD','Azure Commitment Discount: the discount level a P3 must beat to be worth it'),
 ('Deal Desk','the approval path for above-published P3 discounts'),
 ('R1-R7','the seven technical optimization levers detailed in the runbook and in this kit'),
 ('Ramp','program maturation factor in the FY27 forecast, 0 to 100%'),
 ('Cache read','tokens reread from a stable prefix, far cheaper than fresh input'),
 ('Frontier model','top-end model class; the most expensive tokens on the bill'),
]
th(ws,7,2,['Term','Definition'])
for i,(a,b) in enumerate(gl):
    cellv(ws,8+i,2,a,size=10,bold=True); cellv(ws,8+i,3,b,size=9)
    ws.cell(row=8+i,column=3).alignment=Alignment(wrap_text=True)
colw(ws,[2,18,110])
footer(ws,8+len(gl)+1)

# ============================================================= SOURCES
ws=wb.create_sheet('Sources')
tag(ws,'REFERENCE',RED,'Sources and Assumptions','Where every number comes from, and what is editable.')
src=[
 ('BTG billing data','GitHub Token Based Billing report, fiscal month April 2026 (official). Consumption USD 231,704; promo scenario USD 137,624.'),
 ('db.json v2.0.0','Brazil UBB consolidated collection. BTG row: today 130,948; no action 231,704; promo 137,624; licenses 127,249. This kit is coherent with it.'),
 ('Engine reconciliation','The engine computes the promo scenario at ~USD 137.5k from CB 1,693 / CE 2,438. The official report shows 137,624; the USD ~121 gap comes from seat split rounding in the source. Both round to +5% vs today.'),
 ('Monthly FY26 series','db.json uso_mensal, MSX FY26 (Jul/25 to May/26). GHCP ACR grew 4.2x over the series.'),
 ('Model mix','db.json modelos_tokens, Model Level Details April 2026. Sum matches the 231,704 consumption.'),
 ('Plan table','GitHub Copilot list: Business USD 19 and 1,900 credits; Enterprise USD 39 and 3,900; promo 3,000 and 7,000. Verify at docs.github.com before client use.'),
 ('R1-R7 bands','Official mechanics documented in the PT-BR runbook v2.1.0 and the internal HTML v8. Bands are ranges with sources, never promises.'),
 ('Simulator defaults','Lean-start values; combined effect is multiplicative. Mature programs reach 55-70%.'),
 ('P3 and Deal Desk','Subject to eligibility and approval. Not a GitHub announcement; do not present as guaranteed.'),
 ('Promo window','90 days, June to August 2026, expiry 2026-09-01. The September jump is the decision deadline.'),
 ('FY27 ramp curve','Editable assumption reflecting typical program maturation; replace with actuals as they land.'),
]
th(ws,7,2,['Item','Origin / note'])
for i,(a,b) in enumerate(src):
    cellv(ws,8+i,2,a,size=10,bold=True); cellv(ws,8+i,3,b,size=9)
    ws.cell(row=8+i,column=3).alignment=Alignment(wrap_text=True)
    ws.row_dimensions[8+i].height=26
colw(ws,[2,24,116])
footer(ws,8+len(src)+1)

# ============================================================= DASHBOARD (built last; placed after How to Use)
ws=wb.create_sheet('Dashboard',2)
ws.sheet_view.showGridLines=False
c=ws['B2']; c.value='EXECUTIVE DASHBOARD'; c.font=F(bold=True,size=8,color=WHITE); c.fill=HDR(DARK)
ws['B4']='BTG Pactual, UBB Transition at a Glance'; ws['B4'].font=F(bold=True,size=18,color=DARK)
ws['B5']='Live view: every number recalculates from Client Intake, Engine, Simulator and the execution sheets.'
ws['B5'].font=F(size=10,color='666666')
# KPI cards row: label r7, value r8, note r9, color bar r10
cards=[
 ('TODAY (PRU)', "=Engine!C23", CUR, 'per month, audited Apr/26', BLUE),
 ('NO ACTION (SEP+)', "=Engine!C24", CUR, 'per month, +77% vs today', RED),
 ('WITH PROMO (JUN-AUG)', "=Engine!C25", CUR, 'per month, +5% vs today', GREEN),
 ('SEPTEMBER JUMP', "=Engine!C26", CUR, 'per month, the deadline', YELLOW),
 ('MATURE SAVING / YR', "=Simulator!C21", CUR, 'R1-R7 at maturity', GREEN),
 ('FY27 PROGRAM SAVINGS', "='Forecast FY27'!F22", CUR, 'maturity curve', BLUE),
]
for i,(lbl,f,nf,nt,colr) in enumerate(cards):
    c0=2+i*3
    cellv(ws,7,c0,lbl,size=8,bold=True,fontcolor='666666')
    v=cellv(ws,8,c0,f,nf,bold=True,size=15)
    cellv(ws,9,c0,nt,size=8,fontcolor='999999')
    ws.cell(row=10,column=c0).fill=HDR(colr); ws.cell(row=10,column=c0+1).fill=HDR(colr)
# second row of ops KPIs
ops=[
 ('CONSUMPTION / INCLUDED','=Engine!C18','0.00"x"','pool covers '+'',''),
 ('POOL COVERAGE','=Engine!C19',PCT,'of consumption',''),
 ('CLAUDE FAMILY SHARE',f"='Model Mix'!C{MM['total_row']+1}",PCT,'of the token bill',''),
 ('TOP 2 MODELS SHARE',f"='Model Mix'!C{MM['total_row']+2}",PCT,'concentration',''),
 ('ACTIONS DONE',f"=COUNTIF('Action Plan R1-R7'!I{AP['r0']}:I{AP['last']},\"Done\")&\" / \"&COUNTA('Action Plan R1-R7'!B{AP['r0']}:B{AP['last']})",None,'action plan',''),
 ('BLOCKED ITEMS',f"=COUNTIF('Action Plan R1-R7'!I{AP['r0']}:I{AP['last']},\"Blocked\")+COUNTIF('90-Day Plan'!G{WP['r0']}:G{WP['last']},\"Blocked\")",NUM,'needs attention',''),
]
for i,(lbl,f,nf,nt,_) in enumerate(ops):
    c0=2+i*3
    cellv(ws,12,c0,lbl,size=8,bold=True,fontcolor='666666')
    cellv(ws,13,c0,f,nf,bold=True,size=13)
    cellv(ws,14,c0,nt,size=8,fontcolor='999999')
# charts
bar=BarChart(); bar.type='col'; bar.title='The three scenarios (USD/month)'; bar.style=10
data=Reference(wb['Engine'], min_col=3, min_row=23, max_row=25)
cats=Reference(wb['Engine'], min_col=2, min_row=23, max_row=25)
bar.add_data(data, titles_from_data=False); bar.set_categories(cats)
bar.legend=None; bar.height=7.2; bar.width=15
ws.add_chart(bar,'B16')

line=LineChart(); line.title='FY27 overage: no action vs with program (USD/month)'; line.style=12
d1=Reference(wb['Forecast FY27'], min_col=4, min_row=FC['r0']-1+1, max_row=FC['last'])  # header trick below
# use explicit series with headers
line=LineChart(); line.title='FY27 overage: no action vs with program (USD/month)'
s1=Reference(wb['Forecast FY27'], min_col=4, min_row=9, max_row=FC['last'])
s2=Reference(wb['Forecast FY27'], min_col=5, min_row=9, max_row=FC['last'])
catsf=Reference(wb['Forecast FY27'], min_col=2, min_row=FC['r0'], max_row=FC['last'])
# headers at row 7? headers are at row 7 (th). use min_row=7? Series titles from row 7.
s1=Reference(wb['Forecast FY27'], min_col=4, min_row=7, max_row=FC['last'])
s2=Reference(wb['Forecast FY27'], min_col=5, min_row=7, max_row=FC['last'])
line.add_data(s1, titles_from_data=True); line.add_data(s2, titles_from_data=True)
line.set_categories(catsf)
line.height=7.2; line.width=15
ws.add_chart(line,'J16')

mix=BarChart(); mix.type='bar'; mix.title='Model mix, top 8 by token cost (USD, Apr/26)'
dm=Reference(wb['Model Mix'], min_col=3, min_row=MM['r0'], max_row=MM['r0']+7)
cm=Reference(wb['Model Mix'], min_col=2, min_row=MM['r0'], max_row=MM['r0']+7)
mix.add_data(dm, titles_from_data=False); mix.set_categories(cm); mix.legend=None
mix.height=7.2; mix.width=15
ws.add_chart(mix,'B32')

cons=LineChart(); cons.title='FY26 GHCP ACR, audited series (USD/month)'
dc=Reference(wb['Consumption FY26'], min_col=4, min_row=7, max_row=7+len(mensal))
cc=Reference(wb['Consumption FY26'], min_col=2, min_row=8, max_row=7+len(mensal))
cons.add_data(dc, titles_from_data=True); cons.set_categories(cc); cons.legend=None
cons.height=7.2; cons.width=15
ws.add_chart(cons,'J32')
colw(ws,[2]+[13,9,2]*6)
footer(ws,49)

client_slug=CLIENT['name'].split()[0] if 'BTG' not in CLIENT['name'] else 'BTG'
fn=f"Kit_Transicao_UBB_{client_slug}_{CLIENT['version']}_{CLIENT['date']}_EN.xlsx"
wb.save(fn)
print('built ok')
